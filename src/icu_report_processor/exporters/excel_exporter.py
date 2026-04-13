# -*- coding: utf-8 -*-
"""
Excel 导出模块 - 负责将解析结果导出为 Excel 文件
"""

from pathlib import Path
from typing import Dict, Any, List, Optional
from collections import defaultdict

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..parsers.base import ParseResult


class ExcelExporter:
    """
    Excel 导出器
    
    将解析结果整理并导出为结构化的 Excel 文件
    """
    
    def __init__(self, config: Dict[str, Any]):
        """
        初始化导出器
        
        Args:
            config: Excel 配置字典
        """
        self.config = config
        self.fixed_columns = config.get('fixed_columns', [])
        self.sort_column = config.get('sort_column', '主时间')
    
    def export(self, data: Dict[str, List[ParseResult]], 
               output_path: Path,
               patient_info: Optional[Dict[str, str]] = None) -> None:
        """
        导出数据到 Excel
        
        Args:
            data: 按报告类型分组的解析结果字典
            output_path: 输出文件路径
            patient_info: 可选的患者基本信息
        """
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 创建患者信息表（如果提供）
            if patient_info:
                self._create_patient_info_sheet(writer, patient_info)
            
            # 创建各报告类型的工作表
            summary_data = []
            
            for report_type, results in data.items():
                if not results:
                    continue
                
                # 转换为 DataFrame
                df = self._results_to_dataframe(results)
                
                # 排序
                if self.sort_column in df.columns:
                    df = df.sort_values(self.sort_column)
                
                # 写入工作表
                sheet_name = report_type[:31]  # Excel 工作表名最多 31 个字符
                self._create_sheet_with_headers(writer, df, sheet_name, results)
                
                # 记录汇总信息
                summary_data.append({
                    '报告类型': report_type,
                    '记录数': len(results),
                    '时间范围': self._get_time_range(results)
                })
                
                print(f"  已写入工作表: {sheet_name} ({len(results)} 条记录)")
            
            # 创建汇总表
            if summary_data:
                self._create_summary_sheet(writer, summary_data)
        
        print(f"\nExcel 文件已保存: {output_path}")
    
    def _results_to_dataframe(self, results: List[ParseResult]) -> pd.DataFrame:
        """
        将 ParseResult 列表转换为 DataFrame
        
        Args:
            results: 解析结果列表
            
        Returns:
            DataFrame
        """
        records = []
        for result in results:
            records.append(result.to_dict())
        
        df = pd.DataFrame(records)
        
        # 重新排列列顺序
        df = self._reorder_columns(df)
        
        return df
    
    def _reorder_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        重新排列 DataFrame 的列顺序
        
        Args:
            df: 原始 DataFrame
            
        Returns:
            重新排列后的 DataFrame
        """
        # 获取所有列
        all_cols = df.columns.tolist()
        
        # 分离固定列和其他列
        fixed_cols = [c for c in self.fixed_columns if c in all_cols]
        other_cols = [c for c in all_cols if c not in fixed_cols and c != '文件名']
        
        # 新列顺序：固定列 + 检测项目列 + 文件名
        new_cols = fixed_cols + other_cols
        if '文件名' in all_cols:
            new_cols.append('文件名')
        
        # 只保留存在的列
        new_cols = [c for c in new_cols if c in all_cols]
        
        return df[new_cols]
    
    def _create_sheet_with_headers(self, writer: pd.ExcelWriter, 
                                   df: pd.DataFrame, 
                                   sheet_name: str,
                                   results: List[ParseResult]) -> None:
        """
        创建工作表并添加带参考区间的表头
        
        Args:
            writer: ExcelWriter 对象
            df: DataFrame
            sheet_name: 工作表名称
            results: 解析结果列表（用于获取参考区间）
        """
        # 写入数据
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        
        # 获取工作表
        worksheet = writer.sheets[sheet_name]
        
        # 合并参考区间
        reference_ranges = {}
        for result in results:
            reference_ranges.update(result.get_reference_ranges())
        
        # 创建新表头
        headers = df.columns.tolist()
        new_headers = []
        for col in headers:
            if col in reference_ranges and reference_ranges[col]:
                new_headers.append(f"{col}\n(参考: {reference_ranges[col]})")
            else:
                new_headers.append(col)
        
        # 写入新表头
        for col_idx, header in enumerate(new_headers, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.font = Font(bold=True)
        
        # 调整行高
        worksheet.row_dimensions[1].height = 35
        
        # 调整列宽
        self._adjust_column_widths(worksheet, headers, new_headers, df)
    
    def _adjust_column_widths(self, worksheet, headers: List[str], 
                             new_headers: List[str], df: pd.DataFrame) -> None:
        """
        调整列宽以适应内容
        
        Args:
            worksheet: 工作表对象
            headers: 原始表头列表
            new_headers: 新表头列表（含参考区间）
            df: DataFrame
        """
        for col_idx, col in enumerate(headers, 1):
            # 计算列宽
            header_len = len(str(new_headers[col_idx - 1]))
            max_data_len = 0
            
            # 检查该列的数据长度
            if col in df.columns:
                for value in df[col].astype(str):
                    max_data_len = max(max_data_len, len(value))
            
            # 设置列宽
            adjusted_width = min(max(header_len, max_data_len) + 2, 50)
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    def _create_patient_info_sheet(self, writer: pd.ExcelWriter, 
                                   patient_info: Dict[str, str]) -> None:
        """
        创建患者信息工作表
        
        Args:
            writer: ExcelWriter 对象
            patient_info: 患者信息字典
        """
        df = pd.DataFrame([
            {'项目': k, '值': v}
            for k, v in patient_info.items()
        ])
        
        df.to_excel(writer, sheet_name='患者信息', index=False)
        worksheet = writer.sheets['患者信息']
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 20
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter, 
                             summary_data: List[Dict[str, Any]]) -> None:
        """
        创建汇总工作表
        
        Args:
            writer: ExcelWriter 对象
            summary_data: 汇总数据列表
        """
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='数据汇总', index=False)
        
        worksheet = writer.sheets['数据汇总']
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 10
        worksheet.column_dimensions['C'].width = 40
    
    def _get_time_range(self, results: List[ParseResult]) -> str:
        """
        获取结果列表的时间范围
        
        Args:
            results: 解析结果列表
            
        Returns:
            时间范围字符串
        """
        times = [r.main_time for r in results if r.main_time]
        if not times:
            return '未知'
        
        times.sort()
        if len(times) == 1:
            return times[0]
        return f"{times[0]} ~ {times[-1]}"
