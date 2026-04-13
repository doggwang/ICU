# -*- coding: utf-8 -*-
"""
ICU患者指标研究Excel创建脚本 - v3
使用pdfplumber从PDF提取真实数据
患者：王路生，男，67岁，住院号1075642，床号15，ICU
"""

import os
import re
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import pdfplumber
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def extract_text_from_pdf(pdf_path):
    """从PDF中提取所有文本内容"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            return text
    except Exception as e:
        print(f"读取PDF失败 {pdf_path}: {e}")
        return None


def extract_patient_info(text):
    """提取患者基本信息"""
    info = {}

    # 姓名
    name_match = re.search(r'姓\s*名[:\s]*([^\s\n]+)', text)
    if name_match:
        info['姓名'] = name_match.group(1).strip()

    # 病历号/住院号
    id_match = re.search(r'(?:病\s*历\s*号|住\s*院\s*号)[:\s]*([^\s\n]+)', text)
    if id_match:
        info['病历号'] = id_match.group(1).strip()

    # 性别
    gender_match = re.search(r'性\s*别[:\s]*([^\s\n]+)', text)
    if gender_match:
        info['性别'] = gender_match.group(1).strip()

    # 年龄
    age_match = re.search(r'年\s*龄[:\s]*([^\s\n]+)', text)
    if age_match:
        info['年龄'] = age_match.group(1).strip()

    # 科室
    dept_match = re.search(r'科\s*别[:\s]*([^\s\n]+)', text)
    if dept_match:
        info['科室'] = dept_match.group(1).strip()

    # 床号
    bed_match = re.search(r'床\s*号[:\s]*([^\s\n]+)', text)
    if bed_match:
        info['床号'] = bed_match.group(1).strip()

    return info


def extract_time_info(text):
    """从PDF文本中提取时间信息"""
    time_info = {}

    # 匹配采集时间
    collection_match = re.search(r'采集时间[:\s]*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})', text)
    if collection_match:
        time_info['采集时间'] = collection_match.group(1)

    # 匹配接收时间
    receive_match = re.search(r'接收时间[:\s]*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})', text)
    if receive_match:
        time_info['接收时间'] = receive_match.group(1)

    # 匹配报告时间
    report_match = re.search(r'报告时间[:\s]*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})', text)
    if report_match:
        time_info['报告时间'] = report_match.group(1)

    # 获取主时间（优先级：采集时间 > 接收时间 > 报告时间）
    if time_info.get('采集时间'):
        time_info['主时间'] = time_info['采集时间']
    elif time_info.get('接收时间'):
        time_info['主时间'] = time_info['接收时间']
    elif time_info.get('报告时间'):
        time_info['主时间'] = time_info['报告时间']
    else:
        time_info['主时间'] = None

    return time_info


def is_reference_range(value):
    """判断一个值是否是参考区间"""
    if re.match(r'^[\d.]+-[\d.]+$', value):
        return True
    if re.match(r'^[<>][\d.]+$', value):
        return True
    return False


def extract_table_data(text):
    """从文本中提取表格数据"""
    data = {}
    reference_ranges = {}

    lines = text.split('\n')

    for line in lines:
        # 匹配格式：序号 项目名称 结果 [↑↓] 参考区间
        match = re.match(r'\d+\s+\*?([^\d]+?)\s+([\d<>.]+)\s*([\u2191\u2193]?)\s+([^\s]+)', line)
        if match:
            item_name = match.group(1).strip()
            result = match.group(2).strip()
            trend = match.group(3).strip()
            reference = match.group(4).strip()

            item_name = item_name.replace('*', '').strip()

            if trend:
                result = f"{result} {trend}"

            if item_name and result and len(item_name) < 30:
                data[item_name] = result
                if is_reference_range(reference):
                    reference_ranges[item_name] = reference

    return data, reference_ranges


def process_pdf_file(pdf_path):
    """处理单个PDF文件"""
    text = extract_text_from_pdf(str(pdf_path))
    if not text:
        return None, None

    time_info = extract_time_info(text)
    patient_info = extract_patient_info(text)
    table_data, reference_ranges = extract_table_data(text)

    result = {
        '文件名': pdf_path.name,
        **time_info,
        **patient_info,
        **table_data
    }

    return result, reference_ranges


# 定义颜色
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
blue_fill = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_patient_info_sheet(wb):
    """创建患者信息工作表"""
    ws = wb.create_sheet(title="患者信息")
    patient_data = [
        ["姓名", "王路生"],
        ["性别", "男"],
        ["年龄", "67岁"],
        ["住院号", "1075642"],
        ["床号", "15"],
        ["科室", "ICU"],
        ["入院日期", "2026-04-01"],
    ]
    for row_idx, (label, value) in enumerate(patient_data, 1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    return ws


def set_cell_value(ws, row, col, value, fill=None):
    """设置单元格值和样式"""
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    return cell


def create_data_sheet(wb, name, headers, reference_ranges, data):
    """创建数据工作表"""
    ws = wb.create_sheet(title=name)

    # 写入表头
    for col, header in enumerate(headers, 1):
        set_cell_value(ws, 1, col, header, header_fill)

    # 写入参考范围
    for col, ref in enumerate(reference_ranges, 1):
        set_cell_value(ws, 2, col, ref)

    # 写入数据
    for row_idx, row_data in enumerate(data, 3):
        for col_idx, value in enumerate(row_data, 1):
            set_cell_value(ws, row_idx, col_idx, value)

    # 自动调整列宽
    for col in range(1, len(headers) + 1):
        max_length = max(len(str(headers[col-1])), len(str(reference_ranges[col-1])))
        for row in range(3, len(data) + 3):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 40)

    return ws


def main():
    """主函数"""
    base_path = r'c:\Users\39863\Desktop\ICU\检测报告整理'

    print("=" * 60)
    print("ICU患者指标研究Excel创建工具 (v3)")
    print("=" * 60)

    # 存储所有数据
    all_data = defaultdict(list)
    all_references = defaultdict(dict)

    # 定义报告类型和文件夹
    report_folders = {
        '血常规': '血常规',
        '血气分析': '血气分析',
        '生化检验': '生化检验',
        '尿常规': '尿常规',
        'D二聚体': 'D二聚体',
        'BNP': 'BNP心衰标志物',
        'PCT': 'PCT降钙素原',
        'ACT': 'ACT活化凝血时间',
        '心脏超声': '心脏超声',
        'TBNK': 'TBNK免疫细胞',
        '炎症因子': '炎症因子',
        '药敏试验': '药敏试验',
    }

    # 处理每种报告类型
    for sheet_name, folder_name in report_folders.items():
        folder_path = Path(base_path) / folder_name
        if not folder_path.exists():
            print(f"\n文件夹不存在: {folder_path}")
            continue

        pdf_files = list(folder_path.glob('*.pdf'))
        print(f"\n处理 {sheet_name}: 找到 {len(pdf_files)} 个PDF文件")

        for pdf_file in pdf_files:
            data, references = process_pdf_file(pdf_file)
            if data:
                all_data[folder_name].append(data)
                # 合并参考区间
                for item, ref in references.items():
                    if item not in all_references[folder_name]:
                        all_references[folder_name][item] = ref

    # 统计
    total_records = sum(len(records) for records in all_data.values())
    print(f"\n共提取 {total_records} 条记录")

    # 创建Excel
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # 创建患者信息表
    create_patient_info_sheet(wb)

    # 定义工作表名称映射
    sheet_mapping = {
        '血常规': '血常规_原始',
        '血气分析': '血气分析_原始',
        '生化检验': '生化检验_原始',
        '尿常规': '尿常规_原始',
        'D二聚体': 'D二聚体_原始',
        'BNP心衰标志物': 'BNP_原始',
        'PCT降钙素原': 'PCT_原始',
        'ACT活化凝血时间': 'ACT_原始',
        '心脏超声': '心脏超声_原始',
        'TBNK免疫细胞': 'TBNK_原始',
        '炎症因子': '炎症因子_原始',
        '药敏试验': '药敏试验_原始',
    }

    # 为每种报告类型创建工作表
    for folder_name, records in all_data.items():
        if not records:
            continue

        df = pd.DataFrame(records)
        sheet_name = sheet_mapping.get(folder_name, folder_name + '_原始')

        # 确定时间列
        time_col = None
        if '采集时间' in df.columns:
            time_col = '采集时间'
        elif '接收时间' in df.columns:
            time_col = '接收时间'
        elif '报告时间' in df.columns:
            time_col = '报告时间'
        elif '主时间' in df.columns:
            time_col = '主时间'

        # 固定列
        fixed_cols = ['文件名', '主时间', '采集时间', '接收时间', '报告时间', '姓名', '病历号', '性别', '年龄', '科室', '床号']
        other_cols = [c for c in df.columns if c not in fixed_cols and c != '文件名']

        # 重新排列列
        cols = [c for c in fixed_cols if c in df.columns] + other_cols
        cols = [c for c in cols if c in df.columns]
        df = df[cols]

        # 按时间排序
        if time_col and time_col in df.columns:
            df = df.sort_values(time_col, ascending=False)

        print(f"\n{sheet_name}: {len(df)} 条记录")
        print(f"  列: {list(df.columns[:10])}...")

        # 获取参考区间
        ref_dict = all_references.get(folder_name, {})

        # 转换时间列为字符串
        for col in ['采集时间', '接收时间', '报告时间', '主时间']:
            if col in df.columns:
                df[col] = df[col].astype(str)

        # 写入工作表
        headers = df.columns.tolist()
        ref_ranges = [ref_dict.get(col, '') for col in headers]
        data_rows = df.values.tolist()

        create_data_sheet(wb, sheet_name, headers, ref_ranges, data_rows)

    # 保存
    output_path = r'c:\Users\39863\Desktop\ICU\患者指标研究_王路生_v3.xlsx'
    wb.save(output_path)

    print("\n" + "=" * 60)
    print(f"Excel文件已创建: {output_path}")
    print(f"共 {len(wb.sheetnames)} 个工作表:")
    for name in wb.sheetnames:
        print(f"  - {name}")
    print("=" * 60)


if __name__ == '__main__':
    main()