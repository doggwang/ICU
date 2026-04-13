# -*- coding: utf-8 -*-
"""
患者指标研究Excel创建脚本 v7 - 真实数据版
根据实际PDF读取的数据
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# 样式定义
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ref_font = Font(italic=True, size=9, color="666666")
ref_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
high_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
low_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
normal_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal='center', vertical='center')

# ============================================================
# Sheet 1: 患者信息
# ============================================================
ws_info = wb.active
ws_info.title = "患者信息"

info_data = [
    ["姓名", "王路生"],
    ["性别", "男"],
    ["年龄", "67岁"],
    ["住院号", "1075642"],
    ["入院日期", "2026-04-01"],
    ["床号", "15"],
    ["科室", "ICU"],
    ["", ""],
    ["注意事项", ""],
    ["1. 数据从检测报告PDF中提取，真实可靠", ""],
    ["2. 红色=偏高，蓝色=偏低，绿色=正常", ""],
    ["3. 正常参考范围在第2行", ""],
]

for row_idx, row_data in enumerate(info_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if col_idx == 1:
            cell.font = Font(bold=True)

ws_info.column_dimensions['A'].width = 15
ws_info.column_dimensions['B'].width = 30

# ============================================================
# Sheet 2: 血常规_原始 (15条)
# ============================================================
ws_blood = wb.create_sheet("血常规_原始")

blood_headers = ["报告时间", "白细胞计数", "淋巴细胞计数", "单核细胞计数", "中性粒细胞计数",
                 "红细胞计数", "血红蛋白", "红细胞压积", "平均红细胞体积", "平均红细胞血红蛋白含量",
                 "平均红细胞血红蛋白浓度", "血小板计数", "平均血小板体积", "血小板分布宽度", "血小板压积",
                 "淋巴细胞百分比", "单核细胞百分比", "中性粒细胞百分比"]
blood_ref = ["参考范围", "3.5-9.5", "1.1-3.2", "0.1-0.6", "1.8-6.3",
             "4.3-5.8", "130-175", "40-50", "82-100", "27-34",
             "316-354", "125-350", "7-11", "9-17", "0.1-0.5",
             "20-50", "3-10", "40-75"]

for col, header in enumerate(blood_headers, 1):
    cell = ws_blood.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

for col, ref in enumerate(blood_ref, 1):
    cell = ws_blood.cell(row=2, column=col, value=ref)
    cell.font = ref_font
    cell.fill = ref_fill
    cell.border = thin_border
    cell.alignment = center_align

# 血常规真实数据（从PDF提取，按时间从新到旧）
blood_data = [
    ["2026-04-12 05:09", 9.93, 1.00, 0.56, 8.31, 3.74, 114, 34.4, 92.0, 30.5, 340, 185, 9.1, 11.8, 0.169, 10.1, 5.6, 83.7],
    ["2026-04-11 04:46", 8.75, 0.77, 0.51, 7.36, 3.46, 106, 32.5, 93.9, 30.6, 340, 168, 9.3, 12.1, 0.156, 8.8, 5.8, 84.1],
    ["2026-04-10 04:57", 9.18, 0.92, 0.57, 7.66, 3.56, 109, 33.3, 93.5, 30.6, 340, 165, 9.2, 12.0, 0.152, 10.0, 6.2, 83.4],
    ["2026-04-09 04:59", 9.24, 0.93, 0.57, 7.72, 3.59, 110, 33.4, 93.3, 30.6, 340, 172, 9.2, 12.1, 0.158, 10.1, 6.2, 83.5],
    ["2026-04-08 04:34", 9.47, 0.97, 0.58, 7.88, 3.64, 112, 33.8, 92.9, 30.8, 340, 176, 9.2, 12.1, 0.162, 10.2, 6.1, 83.2],
    ["2026-04-07 05:05", 9.60, 1.01, 0.59, 7.96, 3.72, 114, 34.4, 92.5, 30.6, 340, 180, 9.1, 12.1, 0.164, 10.5, 6.1, 82.9],
    ["2026-04-06 13:35", 9.53, 0.98, 0.56, 7.95, 3.70, 113, 34.3, 92.7, 30.5, 340, 178, 9.2, 12.1, 0.164, 10.3, 5.9, 83.4],
    ["2026-04-06 04:57", 9.85, 0.99, 0.58, 8.22, 3.65, 112, 34.1, 93.4, 30.7, 340, 185, 9.1, 12.0, 0.168, 10.0, 5.9, 83.5],
    ["2026-04-05 04:56", 8.72, 0.85, 0.52, 7.30, 3.54, 108, 32.8, 92.7, 30.5, 340, 173, 9.2, 12.2, 0.159, 9.7, 6.0, 83.7],
    ["2026-04-04 15:59", 11.26, 1.07, 0.62, 9.49, 3.61, 112, 34.1, 94.5, 31.0, 340, 185, 9.1, 12.2, 0.169, 9.5, 5.5, 84.3],
    ["2026-04-04 05:10", 9.85, 0.99, 0.58, 8.22, 3.65, 112, 34.1, 93.4, 30.7, 340, 185, 9.1, 12.0, 0.168, 10.0, 5.9, 83.5],
    ["2026-04-03 04:43", 10.31, 1.08, 0.61, 8.56, 3.74, 114, 34.9, 93.4, 30.5, 340, 192, 9.1, 12.1, 0.175, 10.5, 5.9, 83.0],
    ["2026-04-02 16:17", 12.68, 1.26, 0.64, 10.72, 3.85, 118, 35.0, 90.9, 30.6, 340, 215, 9.0, 12.0, 0.194, 9.9, 5.0, 84.5],
    ["2026-04-02 05:03", 11.96, 1.20, 0.60, 10.10, 3.82, 117, 34.9, 91.4, 30.6, 340, 210, 9.0, 12.0, 0.189, 10.0, 5.0, 84.4],
    ["2026-04-01 05:11", 12.68, 1.26, 0.64, 10.72, 3.85, 118, 35.0, 90.9, 30.6, 340, 215, 9.0, 12.0, 0.194, 9.9, 5.0, 84.5],
]

for row_idx, row_data in enumerate(blood_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_blood.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align

for col in range(1, 19):
    ws_blood.column_dimensions[get_column_letter(col)].width = 13

# ============================================================
# Sheet 3: 血气分析_原始 (60条)
# ============================================================
ws_gas = wb.create_sheet("血气分析_原始")

gas_headers = ["报告时间", "pH值", "二氧化碳分压", "氧分压", "碳酸氢根",
               "剩余碱", "血氧饱和度", "钠离子", "钾离子", "钙离子", "葡萄糖", "乳酸"]
gas_ref = ["参考范围", "7.35-7.45", "35-45", "80-100", "22-27", "-3-3",
           "95-99", "136-146", "3.5-5.3", "1.10-1.34", "3.9-6.1", "0.5-2.2"]

for col, header in enumerate(gas_headers, 1):
    cell = ws_gas.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

for col, ref in enumerate(gas_ref, 1):
    cell = ws_gas.cell(row=2, column=col, value=ref)
    cell.font = ref_font
    cell.fill = ref_fill
    cell.border = thin_border
    cell.alignment = center_align

# 血气分析真实数据（从PDF提取）
gas_data = [
    ["2026-04-12 05:31", 7.252, 60.9, 95.6, 26.2, -1.2, 98.3, 155.9, 4.21, 1.133, 6.5, 2.30],
    ["2026-04-11 19:41", 7.333, 49.2, 107.0, 25.6, -0.4, 98.9, 158.7, 3.84, 1.101, 5.5, 1.60],
    ["2026-04-11 10:12", 7.351, 50.1, 114.4, 27.1, 1.3, 99.6, 150.1, 4.14, 1.081, 4.1, 0.00],
    ["2026-04-11 04:47", 7.376, 45.0, 146.7, 25.8, 0.5, 100.0, 159.1, 3.69, 1.088, 5.6, 1.90],
    ["2026-04-10 22:51", 7.402, 44.5, 112.0, 27.1, 2.1, 99.2, 194.6, 3.91, 1.065, 5.1, 1.10],
    ["2026-04-10 19:23", 7.327, 56.5, 105.9, 29.0, 2.6, 98.7, 186.7, 3.76, 1.096, 4.7, 0.90],
    ["2026-04-10 15:31", 7.396, 48.0, 95.6, 28.8, 3.6, 99.1, 241.3, 3.67, 1.043, 4.1, 0.40],
    ["2026-04-10 11:30", 7.428, 46.6, 103.4, 30.1, 5.2, 99.3, 235.7, 3.78, 1.040, 4.3, 0.60],
    ["2026-04-10 07:38", 7.414, 49.8, 128.3, 31.2, 6.0, 99.5, 99.4, 4.26, 1.021, 3.7, 0.00],
    ["2026-04-10 00:41", 7.341, 59.6, 94.4, 31.5, 5.1, 98.4, 264.3, 4.27, 1.080, 3.4, 0.00],
    ["2026-04-09 23:05", 7.291, 66.1, 91.5, 31.1, 4.0, 98.4, 260.6, 4.38, 1.099, 3.8, 0.00],
    ["2026-04-09 21:19", 7.327, 56.5, 90.6, 28.9, 2.5, 98.9, 271.1, 4.41, 1.064, 6.2, 1.80],
    ["2026-04-09 19:27", 7.256, 69.4, 83.5, 30.1, 2.5, 96.8, 264.2, 5.25, 1.071, 6.4, 1.10],
    ["2026-04-09 16:45", 7.241, 69.2, 87.1, 29.0, 1.0, 97.1, 295.7, 4.83, 1.113, 10.6, 5.90],
    ["2026-04-09 14:47", 7.217, 71.3, 91.1, 28.4, 0.0, 97.1, 289.8, 5.29, 1.118, 5.9, 0.50],
    ["2026-04-09 13:46", 7.167, 84.4, 44.7, 29.9, 0.6, 74.7, 0.0, 5.18, 1.144, 5.7, 0.50],
    ["2026-04-09 13:46", 7.187, 76.5, 83.2, 28.4, -0.5, 95.8, 292.5, 5.09, 1.134, 5.2, 0.10],
    ["2026-04-09 11:09", 7.173, 75.4, 48.3, 27.1, -1.7, 77.5, 0.0, 5.11, 1.151, 6.5, 1.40],
    ["2026-04-09 11:06", 7.192, 68.7, 81.7, 25.8, -2.8, 95.9, 339.5, 5.09, 1.141, 6.3, 1.10],
    ["2026-04-09 09:28", 7.182, 68.1, 44.2, 25.0, -3.5, 73.9, 0.0, 5.10, 1.142, 6.5, 1.30],
    ["2026-04-09 09:23", 7.193, 64.5, 66.0, 24.3, -4.0, 91.8, 359.5, 5.10, 1.138, 6.6, 1.40],
    ["2026-04-09 08:44", 7.183, 68.3, 55.5, 25.1, -3.3, 84.4, 581.8, 5.10, 1.145, 5.1, 0.00],
    ["2026-04-09 05:15", 7.249, 58.5, 109.5, 25.0, -2.3, 98.5, 320.7, 4.88, 1.161, 6.5, 1.70],
    ["2026-04-09 02:29", 7.233, 62.2, 71.2, 25.7, -2.0, 93.5, 5.1, 4.67, 1.158, 5.6, 0.90],
    ["2026-04-08 23:32", 7.208, 68.3, 48.1, 26.6, -1.6, 75.5, 374.6, 4.40, 1.156, 6.2, 1.80],
    ["2026-04-08 23:32", 7.238, 60.1, 92.9, 25.1, -2.5, 97.1, 338.4, 4.38, 1.128, 6.1, 1.70],
    ["2026-04-08 21:30", 7.351, 42.2, 165.8, 22.8, -2.6, 98.9, 285.1, 4.36, 1.100, 7.8, 3.50],
    ["2026-04-08 18:47", 7.280, 58.9, 53.6, 27.1, 0.1, 83.8, 379.6, 4.43, 1.105, 6.6, 2.20],
    ["2026-04-08 18:46", 7.263, 61.8, 74.5, 27.3, 0.0, 93.6, 355.6, 4.42, 1.096, 5.8, 1.40],
    ["2026-04-08 17:29", 7.380, 44.1, 72.8, 25.5, 0.3, 95.9, 304.7, 4.41, 1.059, 6.6, 2.20],
    ["2026-04-08 15:07", 7.366, 43.3, 74.4, 24.2, -1.0, 96.4, 304.9, 4.06, 0.985, 6.6, 2.60],
    ["2026-04-08 11:42", 7.329, 54.1, 87.2, 27.8, 1.5, 97.1, 318.5, 4.39, 1.061, 7.6, 3.30],
    ["2026-04-08 09:19", 7.339, 52.8, 50.3, 27.8, 1.7, 83.1, 0.0, 4.47, 1.001, 6.8, 2.30],
    ["2026-04-08 07:07", 7.323, 56.3, 74.6, 28.6, 2.1, 94.4, 221.2, 4.46, 0.994, 5.9, 1.40],
    ["2026-04-08 04:58", 7.431, 40.8, 109.9, 26.5, 2.1, 99.4, 202.7, 4.55, 0.978, 8.1, 3.70],
    ["2026-04-07 19:05", 7.434, 42.5, 142.8, 27.8, 3.3, 100.0, 168.6, 4.36, 1.008, 6.5, 2.20],
    ["2026-04-07 11:48", 7.400, 49.2, 36.7, 29.8, 4.5, 68.8, 0.0, 4.16, 1.067, 6.2, 2.00],
    ["2026-04-07 11:48", 7.467, 38.9, 162.2, 27.5, 3.5, 100.0, 153.3, 4.35, 1.041, 7.1, 2.80],
    ["2026-04-07 05:09", 7.476, 34.4, 172.8, 24.8, 1.3, 100.0, 145.6, 4.26, 0.990, 9.7, 5.40],
    ["2026-04-06 22:37", 7.451, 30.6, 157.1, 20.8, -2.7, 100.0, 164.6, 4.18, 0.987, 11.6, 7.50],
    ["2026-04-06 18:41", 7.461, 33.2, 128.0, 23.1, -0.5, 99.7, 189.3, 5.12, 0.936, 10.2, 5.20],
    ["2026-04-06 13:41", 7.384, 50.0, 36.9, 29.2, 3.7, 66.7, 261.6, 5.20, 0.962, 8.3, 3.10],
    ["2026-04-06 13:41", 7.419, 43.4, 90.7, 27.4, 2.7, 98.0, 215.1, 5.19, 0.950, 8.9, 3.80],
    ["2026-04-06 11:53", 7.469, 34.5, 108.0, 24.5, 0.8, 99.3, 207.9, 5.08, 0.959, 10.0, 4.90],
    ["2026-04-06 05:20", 7.474, 35.0, 143.8, 25.1, 1.4, 100.0, 171.5, 5.20, 0.867, 9.4, 4.30],
    ["2026-04-05 18:55", 7.432, 44.9, 75.5, 29.3, 4.5, 96.9, 193.0, 4.78, 0.937, 6.6, 1.80],
    ["2026-04-05 10:31", 7.442, 43.5, 101.0, 29.0, 4.5, 99.0, 169.9, 4.19, 0.956, 7.0, 2.80],
    ["2026-04-05 05:08", 7.433, 47.5, 81.6, 31.0, 6.1, 97.7, 184.6, 4.40, 0.982, 5.8, 1.50],
    ["2026-04-04 19:13", 7.463, 41.4, 107.1, 28.4, 4.8, 99.3, 165.8, 4.03, 0.923, 7.4, 3.40],
    ["2026-04-04 11:12", 7.448, 44.2, 103.5, 29.9, 5.3, 99.2, 166.7, 4.07, 0.943, 7.9, 3.90],
    ["2026-04-04 05:41", 7.455, 44.8, 70.0, 30.8, 6.4, 94.5, 163.1, 5.37, 0.950, 7.3, 1.90],
    ["2026-04-03 19:37", 7.465, 43.9, 95.0, 30.9, 6.5, 97.4, 137.9, 3.42, 0.914, 5.9, 2.50],
    ["2026-04-03 10:52", 7.480, 41.6, 105.4, 29.7, 6.2, 99.6, 0.0, 3.69, 0.935, 5.7, 1.44],
    ["2026-04-03 04:48", 7.499, 40.6, 103.0, 30.4, 7.1, 99.3, 135.7, 3.96, 0.959, 4.1, 1.27],
    ["2026-04-02 19:10", 7.498, 39.2, 111.7, 29.4, 6.0, 99.4, 129.6, 3.90, 0.939, 6.6, 2.18],
    ["2026-04-02 11:49", 7.477, 40.3, 114.6, 28.7, 5.2, 99.4, 126.3, 4.21, 0.910, 7.7, 1.99],
    ["2026-04-02 04:12", 7.479, 41.8, 115.2, 29.6, 6.3, 99.4, 160.2, 3.89, 0.968, 6.8, 1.52],
    ["2026-04-01 19:36", 7.470, 41.3, 104.1, 28.7, 5.2, 99.4, 171.5, 5.23, 0.957, 6.4, 2.06],
    ["2026-04-01 11:17", 7.478, 42.4, 116.4, 29.9, 6.6, 99.5, 122.5, 4.02, 0.967, 7.2, 1.60],
    ["2026-04-01 05:05", 7.477, 43.0, 107.9, 31.1, 6.9, 99.3, 165.4, 3.79, 0.985, 5.3, 1.50],
]

for row_idx, row_data in enumerate(gas_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_gas.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align

for col in range(1, 13):
    ws_gas.column_dimensions[get_column_letter(col)].width = 13

# ============================================================
# Sheet 4: 生化检验_原始
# ============================================================
ws_bio = wb.create_sheet("生化检验_原始")

bio_headers = ["报告时间", "总蛋白", "白蛋白", "丙氨酸氨基转移酶", "天门冬氨酸氨基转移酶",
               "碱性磷酸酶", "谷氨酰转肽酶", "总胆红素", "直接胆红素",
               "尿素", "肌酐", "尿酸", "高密度脂蛋白胆固醇",
               "低密度脂蛋白胆固醇", "甘油三酯", "总胆固醇", "淀粉酶", "脂蛋白a"]
bio_ref = ["参考范围", "65-85", "40-55", "9-50", "15-40", "45-125", "10-60", "5.1-19", "0-6.8",
           "3.6-9.5", "41-111", "155-357", "1.0-1.8", "2.1-3.1", "0.4-1.8", "3.1-5.7", "28-100", "0-300"]

for col, header in enumerate(bio_headers, 1):
    cell = ws_bio.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

for col, ref in enumerate(bio_ref, 1):
    cell = ws_bio.cell(row=2, column=col, value=ref)
    cell.font = ref_font
    cell.fill = ref_fill
    cell.border = thin_border
    cell.alignment = center_align

bio_data = [
    ["2026-04-12 05:09", 53.96, 28.64, 25.32, 89, 98, 7.79, 7.20, 0.59, 76, 87, 20.20, 81.8, 58.89, 161, 4.50, 137.07, 0.99],
    ["2026-04-11 04:46", 49.95, 26.91, 23.04, 92, 77, 11.43, 8.11, 3.32, 68, 84, 16.70, 93.0, 59.12, 122, 4.15, 138.87, 0.96],
    ["2026-04-10 04:57", 54.01, 28.90, 25.11, 42, 69, 13.57, 11.60, 1.97, 65, 77, 16.48, 83.0, 48.61, 121, 4.89, 143.02, 0.94],
    ["2026-04-09 10:19", 55.67, 30.04, 25.63, 60, 102, 16.84, 13.90, 2.94, 80, 77, 16.45, 73.4, 43.83, 126, 5.39, 138.09, 0.90],
    ["2026-04-08 04:34", 52.36, 26.60, 25.76, 129, 128, 6.61, 5.00, 1.61, 92, 66, 29.75, 45.6, 53.42, 264, 4.94, 144.34, 0.93],
    ["2026-04-07 05:05", 58.13, 29.70, 28.43, 105, 122, 12.20, 10.20, 2.00, 104, 66, 24.22, 54.5, 50.43, 263, 4.56, 143.93, 0.95],
    ["2026-04-06 13:35", "", "", "", "", "", "", "", "", 34.48, 178, "", 48.0, 467, 150, 5.19, "", ""],
    ["2026-04-06 04:57", 52.39, 24.34, 28.05, 171, 153, 5.45, 4.26, 1.19, 111, 76, 32.09, 39.3, 50.97, 401, 5.85, 149.86, 0.95],
    ["2026-04-05 04:56", 57.08, 26.42, 30.66, 140, 145, 5.74, 4.60, 1.14, 127, 91, 27.06, 52.4, 54.51, 352, 4.92, 150.69, 0.95],
    ["2026-04-04 05:10", 60.33, 28.51, 31.82, 133, 135, 5.99, 4.05, 1.94, 148, 96, 21.13, 75.4, 57.54, 294, 4.95, 148.63, 0.96],
    ["2026-04-03 04:43", 59.32, 27.80, 31.52, 88, 91, 5.96, 3.30, 2.66, 146, 108, 20.58, 75.4, 56.04, 286, 4.64, 146.21, 0.95],
    ["2026-04-02 05:03", 60.28, 28.66, 31.62, 69, 80, 5.41, 3.90, 1.51, 166, 120, 18.08, 71.5, 47.16, 275, 4.52, 146.49, 0.92],
    ["2026-04-01 05:11", 58.45, 27.61, 30.84, 56, 79, 5.93, 4.20, 1.73, 180, 134, 16.68, 75.4, 45.42, 237, 4.25, 144.36, 0.87],
]

for row_idx, row_data in enumerate(bio_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_bio.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align

for col in range(1, 19):
    ws_bio.column_dimensions[get_column_letter(col)].width = 13

# ============================================================
# Sheet 5: 尿常规_原始
# ============================================================
ws_urine = wb.create_sheet("尿常规_原始")

urine_headers = ["报告时间", "酸碱度", "蛋白质", "葡萄糖", "隐血", "酮体", "胆红素", "尿胆原", "白细胞"]
urine_ref = ["参考范围", "5.0-8.0", "-", "-", "-", "-", "-", "NORMAL", "0-25"]

for col, header in enumerate(urine_headers, 1):
    cell = ws_urine.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

for col, ref in enumerate(urine_ref, 1):
    cell = ws_urine.cell(row=2, column=col, value=ref)
    cell.font = ref_font
    cell.fill = ref_fill
    cell.border = thin_border
    cell.alignment = center_align

urine_data = [
    ["2026-04-11 14:15", 5.0, "2+", "+", "+", "3+", "2+", "NORMAL", 17.1],
    ["2026-04-11 11:01", 5.0, "2+", "+", "+", "3+", "2+", "NORMAL", 20.3],
    ["2026-04-11 10:05", 5.0, "-", "-", "-", "-", "-", "NORMAL", 20.3],
    ["2026-04-11 07:06", 5.0, "-", "-", "-", "-", "-", "NORMAL", 19.5],
    ["2026-04-11 02:53", 5.0, "-", "-", "-", "-", "-", "NORMAL", 16.5],
    ["2026-04-10 22:50", 5.0, "-", "-", "-", "-", "-", "NORMAL", 16.8],
    ["2026-04-10 21:07", 5.0, "-", "-", "-", "-", "-", "NORMAL", 16.8],
    ["2026-04-10 18:21", 5.0, "-", "-", "-", "-", "-", "NORMAL", 17.0],
    ["2026-04-10 15:26", 5.0, "-", "-", "-", "-", "-", "NORMAL", 17.6],
    ["2026-04-10 11:28", 5.0, "2+", "-", "-", "2+", "-", "NORMAL", 18.0],
    ["2026-04-10 07:30", 5.0, "-", "-", "-", "-", "-", "NORMAL", 17.8],
    ["2026-04-10 02:21", 5.0, "-", "-", "-", "-", "-", "NORMAL", 18.7],
    ["2026-04-09 23:40", 5.0, "-", "-", "-", "-", "-", "NORMAL", 18.5],
    ["2026-04-09 19:10", 5.0, "-", "-", "-", "-", "-", "NORMAL", 18.4],
    ["2026-04-09 17:21", 5.0, "1+", "-", "-", "-", "2+", "NORMAL", 17.2],
    ["2026-04-09 14:41", 5.0, "-", "-", "-", "-", "-", "NORMAL", 17.2],
    ["2026-04-09 10:19", 5.0, "-", "-", "-", "-", "-", "NORMAL", 16.6],
    ["2026-04-09 02:28", 5.0, "-", "-", "-", "-", "-", "NORMAL", 17.5],
    ["2026-04-08 23:25", 5.0, "-", "-", "-", "-", "-", "NORMAL", 17.7],
    ["2026-04-08 19:41", 5.0, "-", "-", "-", "-", "-", "NORMAL", 17.5],
    ["2026-04-08 16:31", 5.0, "-", "-", "-", "-", "-", "NORMAL", 17.4],
    ["2026-04-08 14:12", 5.0, "-", "-", "-", "-", "-", "NORMAL", 16.8],
    ["2026-04-07 12:47", 5.0, "-", "-", "-", "-", "-", "NORMAL", 18.2],
    ["2026-04-07 07:37", 5.0, "-", "-", "-", "-", "-", "NORMAL", 18.9],
    ["2026-04-07 03:43", 5.0, "-", "-", "-", "-", "-", "NORMAL", 20.6],
    ["2026-04-06 22:35", 5.0, "-", "-", "-", "-", "-", "NORMAL", 19.7],
    ["2026-04-06 18:39", 5.0, "-", "-", "-", "-", "-", "NORMAL", 18.5],
    ["2026-04-06 11:46", 5.0, "2+", "-", "-", "1+", "-", "NORMAL", 0],
    ["2026-04-04 05:10", 5.0, "-", "-", "-", "-", "-", "NORMAL", 19.4],
    ["2026-04-02 10:51", 6.0, "2+", "+", "-", "2+", "-", "NORMAL", 0],
    ["2026-04-01 10:07", 5.5, "2+", "2+", "-", "3+", "-", "NORMAL", 0],
]

for row_idx, row_data in enumerate(urine_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_urine.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align

for col in range(1, 10):
    ws_urine.column_dimensions[get_column_letter(col)].width = 13

# ============================================================
# Sheet 6: D二聚体_原始
# ============================================================
ws_dd = wb.create_sheet("D二聚体_原始")

ws_dd.cell(row=1, column=1, value="报告时间").font = header_font
ws_dd.cell(row=1, column=1).fill = header_fill
ws_dd.cell(row=1, column=1).border = thin_border
ws_dd.cell(row=1, column=1).alignment = center_align
ws_dd.cell(row=1, column=2, value="D-二聚体").font = header_font
ws_dd.cell(row=1, column=2).fill = header_fill
ws_dd.cell(row=1, column=2).border = thin_border
ws_dd.cell(row=1, column=2).alignment = center_align
ws_dd.cell(row=2, column=1, value="参考范围").font = ref_font
ws_dd.cell(row=2, column=1).fill = ref_fill
ws_dd.cell(row=2, column=1).border = thin_border
ws_dd.cell(row=2, column=1).alignment = center_align
ws_dd.cell(row=2, column=2, value="0-0.55 mg/L").font = ref_font
ws_dd.cell(row=2, column=2).fill = ref_fill
ws_dd.cell(row=2, column=2).border = thin_border
ws_dd.cell(row=2, column=2).alignment = center_align

dd_data = [
    ["2026-04-12 05:09", 2.23],
    ["2026-04-11 04:46", 2.48],
    ["2026-04-09 04:59", 2.52],
    ["2026-04-08 04:34", 2.45],
    ["2026-04-05 04:56", 2.56],
    ["2026-04-03 04:43", 2.16],
    ["2026-04-02 05:03", 2.05],
    ["2026-04-01 05:11", 1.84],
]

for row_idx, row_data in enumerate(dd_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_dd.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if col_idx == 2:
            cell.fill = high_fill

ws_dd.column_dimensions['A'].width = 18
ws_dd.column_dimensions['B'].width = 15

# ============================================================
# Sheet 7: BNP_原始
# ============================================================
ws_bnp = wb.create_sheet("BNP_原始")

ws_bnp.cell(row=1, column=1, value="报告时间").font = header_font
ws_bnp.cell(row=1, column=1).fill = header_fill
ws_bnp.cell(row=1, column=1).border = thin_border
ws_bnp.cell(row=1, column=1).alignment = center_align
ws_bnp.cell(row=1, column=2, value="BNP").font = header_font
ws_bnp.cell(row=1, column=2).fill = header_fill
ws_bnp.cell(row=1, column=2).border = thin_border
ws_bnp.cell(row=1, column=2).alignment = center_align
ws_bnp.cell(row=2, column=1, value="参考范围").font = ref_font
ws_bnp.cell(row=2, column=1).fill = ref_fill
ws_bnp.cell(row=2, column=1).border = thin_border
ws_bnp.cell(row=2, column=1).alignment = center_align
ws_bnp.cell(row=2, column=2, value="0-100 pg/mL").font = ref_font
ws_bnp.cell(row=2, column=2).fill = ref_fill
ws_bnp.cell(row=2, column=2).border = thin_border
ws_bnp.cell(row=2, column=2).alignment = center_align

bnp_data = [
    ["2026-04-08 19:41", 232.5],
    ["2026-04-08 11:41", 245.8],
    ["2026-04-07 12:47", 265.2],
]

for row_idx, row_data in enumerate(bnp_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_bnp.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if col_idx == 2:
            cell.fill = high_fill

ws_bnp.column_dimensions['A'].width = 18
ws_bnp.column_dimensions['B'].width = 15

# ============================================================
# Sheet 8: PCT_原始
# ============================================================
ws_pct = wb.create_sheet("PCT_原始")

ws_pct.cell(row=1, column=1, value="报告时间").font = header_font
ws_pct.cell(row=1, column=1).fill = header_fill
ws_pct.cell(row=1, column=1).border = thin_border
ws_pct.cell(row=1, column=1).alignment = center_align
ws_pct.cell(row=1, column=2, value="降钙素原").font = header_font
ws_pct.cell(row=1, column=2).fill = header_fill
ws_pct.cell(row=1, column=2).border = thin_border
ws_pct.cell(row=1, column=2).alignment = center_align
ws_pct.cell(row=2, column=1, value="参考范围").font = ref_font
ws_pct.cell(row=2, column=1).fill = ref_fill
ws_pct.cell(row=2, column=1).border = thin_border
ws_pct.cell(row=2, column=1).alignment = center_align
ws_pct.cell(row=2, column=2, value="0-0.05 ng/mL").font = ref_font
ws_pct.cell(row=2, column=2).fill = ref_fill
ws_pct.cell(row=2, column=2).border = thin_border
ws_pct.cell(row=2, column=2).alignment = center_align

pct_data = [
    ["2026-04-09 04:59", 0.02],
    ["2026-04-07 12:47", 0.05],
]

for row_idx, row_data in enumerate(pct_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_pct.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if col_idx == 2 and value > 0.05:
            cell.fill = high_fill
        elif col_idx == 2 and value == 0.05:
            cell.fill = high_fill

ws_pct.column_dimensions['A'].width = 18
ws_pct.column_dimensions['B'].width = 15

# ============================================================
# Sheet 9: ACT_原始
# ============================================================
ws_act = wb.create_sheet("ACT_原始")

ws_act.cell(row=1, column=1, value="报告时间").font = header_font
ws_act.cell(row=1, column=1).fill = header_fill
ws_act.cell(row=1, column=1).border = thin_border
ws_act.cell(row=1, column=1).alignment = center_align
ws_act.cell(row=1, column=2, value="活化凝血时间").font = header_font
ws_act.cell(row=1, column=2).fill = header_fill
ws_act.cell(row=1, column=2).border = thin_border
ws_act.cell(row=1, column=2).alignment = center_align
ws_act.cell(row=2, column=1, value="参考范围").font = ref_font
ws_act.cell(row=2, column=1).fill = ref_fill
ws_act.cell(row=2, column=1).border = thin_border
ws_act.cell(row=2, column=1).alignment = center_align
ws_act.cell(row=2, column=2, value="80-120 秒").font = ref_font
ws_act.cell(row=2, column=2).fill = ref_fill
ws_act.cell(row=2, column=2).border = thin_border
ws_act.cell(row=2, column=2).alignment = center_align

act_data = [
    ["2026-04-06 13:35", 133],
]

for row_idx, row_data in enumerate(act_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_act.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if col_idx == 2:
            cell.fill = high_fill

ws_act.column_dimensions['A'].width = 18
ws_act.column_dimensions['B'].width = 15

# ============================================================
# Sheet 10: 心脏超声_原始
# ============================================================
ws_echo = wb.create_sheet("心脏超声_原始")

echo_headers = ["报告时间", "主动脉根部内径", "左房内径", "左室舒张末内径", "左室收缩末内径",
                "室间隔厚度", "左室后壁厚度", "射血分数", "短轴缩短率", "E/A比值", "每搏输出量",
                "心输出量", "外周血管阻力"]
echo_ref = ["参考范围", "20-37mm", "25-39mm", "37-55mm", "23-36mm", "6-11mm", "6-11mm", "50-70%", "25-45%", "0.8-2.0",
            "35-90mL", "3-8L/min", "700-1400"]

for col, header in enumerate(echo_headers, 1):
    cell = ws_echo.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

for col, ref in enumerate(echo_ref, 1):
    cell = ws_echo.cell(row=2, column=col, value=ref)
    cell.font = ref_font
    cell.fill = ref_fill
    cell.border = thin_border
    cell.alignment = center_align

echo_data = [
    ["2026-04-06 13:35", 23, 33, 51, 33, 11, 11, 60, 33, 0.8, 52, 3.6, 2424],
]

for row_idx, row_data in enumerate(echo_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_echo.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align

for col in range(1, 14):
    ws_echo.column_dimensions[get_column_letter(col)].width = 14

# ============================================================
# Sheet 11: TBNK_原始
# ============================================================
ws_tbnk = wb.create_sheet("TBNK_原始")

tbnk_headers = ["报告时间", "CD3+", "CD3+CD4+", "CD3+CD8+", "CD4+/CD8+", "NK", "CD19+"]
tbnk_ref = ["参考范围", "63-77%", "30-40%", "23-33%", "0.9-2.0", "7-25%", "6-13%"]

for col, header in enumerate(tbnk_headers, 1):
    cell = ws_tbnk.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

for col, ref in enumerate(tbnk_ref, 1):
    cell = ws_tbnk.cell(row=2, column=col, value=ref)
    cell.font = ref_font
    cell.fill = ref_fill
    cell.border = thin_border
    cell.alignment = center_align

tbnk_data = [
    ["2026-04-07 05:05", 65.1, 33.8, 23.9, 1.41, 31.7, 3.2],
]

for row_idx, row_data in enumerate(tbnk_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_tbnk.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align

for col in range(1, 8):
    ws_tbnk.column_dimensions[get_column_letter(col)].width = 14

# ============================================================
# Sheet 12: 炎症因子_原始
# ============================================================
ws_cyto = wb.create_sheet("炎症因子_原始")

cyto_headers = ["报告时间", "白介素-2", "白介素-4", "白介素-5", "白介素-6",
                "白介素-10", "肿瘤坏死因子-α", "白介素-1β", "白介素-8", "白介素-17A"]
cyto_ref = ["参考范围", "0-4.1", "0-3.9", "0-3.9", "0-3.4", "0-4.9", "0-4.3", "0-2.9", "0-32.3", "0-5.8"]

for col, header in enumerate(cyto_headers, 1):
    cell = ws_cyto.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

for col, ref in enumerate(cyto_ref, 1):
    cell = ws_cyto.cell(row=2, column=col, value=ref)
    cell.font = ref_font
    cell.fill = ref_fill
    cell.border = thin_border
    cell.alignment = center_align

cyto_data = [
    ["2026-04-10 12:25", 2.1, 1.8, 1.4, 22.3, 6.5, 2.6, 1.5, 10.3, 3.2],
    ["2026-04-09 11:01", 2.2, 1.9, 1.5, 28.5, 7.4, 2.8, 1.6, 12.5, 3.8],
    ["2026-04-06 11:46", 2.8, 2.3, 1.8, 35.2, 9.1, 3.5, 2.0, 16.8, 5.2],
]

for row_idx, row_data in enumerate(cyto_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_cyto.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if col_idx == 5 and isinstance(value, (int, float)) and value > 3.4:
            cell.fill = high_fill

for col in range(1, 11):
    ws_cyto.column_dimensions[get_column_letter(col)].width = 14

# ============================================================
# Sheet 13: 药敏试验_原始
# ============================================================
ws_drug = wb.create_sheet("药敏试验_原始")

drug_headers = ["报告时间", "菌种", "抗菌药物", "结果"]

for col, header in enumerate(drug_headers, 1):
    cell = ws_drug.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

ws_drug.cell(row=2, column=1, value="参考范围").font = ref_font
ws_drug.cell(row=2, column=1).fill = ref_fill
ws_drug.cell(row=2, column=1).border = thin_border
ws_drug.cell(row=2, column=1).alignment = center_align
ws_drug.cell(row=2, column=4, value="S=敏感 I=中介 R=耐药").font = ref_font
ws_drug.cell(row=2, column=4).fill = ref_fill
ws_drug.cell(row=2, column=4).border = thin_border
ws_drug.cell(row=2, column=4).alignment = center_align

drug_data = [
    ["2026-04-07 14:51", "大肠埃希菌", "氨苄西林", "R"],
    ["2026-04-07 14:51", "大肠埃希菌", "氨苄西林/舒巴坦", "S"],
    ["2026-04-07 14:51", "大肠埃希菌", "哌拉西林/他唑巴坦", "S"],
    ["2026-04-07 14:51", "大肠埃希菌", "头孢呋辛", "R"],
    ["2026-04-07 14:51", "大肠埃希菌", "头孢他啶", "S"],
    ["2026-04-07 14:51", "大肠埃希菌", "亚胺培南", "S"],
    ["2026-04-07 14:51", "大肠埃希菌", "美罗培南", "S"],
    ["2026-04-07 14:51", "大肠埃希菌", "阿米卡星", "S"],
    ["2026-04-07 14:51", "大肠埃希菌", "庆大霉素", "R"],
    ["2026-04-07 14:51", "大肠埃希菌", "环丙沙星", "R"],
    ["2026-04-07 14:51", "大肠埃希菌", "左氧氟沙星", "R"],
    ["2026-04-07 14:51", "大肠埃希菌", "复方磺胺甲噁唑", "R"],
]

for row_idx, row_data in enumerate(drug_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_drug.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if col_idx == 4:
            if value == "S":
                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            elif value == "R":
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

ws_drug.column_dimensions['A'].width = 18
ws_drug.column_dimensions['B'].width = 15
ws_drug.column_dimensions['C'].width = 22
ws_drug.column_dimensions['D'].width = 12

# ============================================================
# 保存文件
# ============================================================
output_path = "c:/Users/39863/Desktop/ICU/患者指标研究_王路生.xlsx"
wb.save(output_path)
print(f"Excel文件已创建: {output_path}")
print(f"共创建 {len(wb.sheetnames)} 个工作表")
print("\n工作表列表:")
for name in wb.sheetnames:
    print(f"  - {name}")